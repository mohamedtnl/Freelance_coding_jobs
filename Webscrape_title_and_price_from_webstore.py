import openpyxl, requests, bs4

wb = openpyxl.Workbook()
sheet = wb['Sheet']         #open new empty excel sheet

                            #web store with products to be scraped
w1 = 'https://www.alternate.de/Spielzeug/Top-Spielzeug-Marken/html/listings/1458038895052/1459763282981/1460961290525/1460962960774/1460963804585/1460966969464/1460967387753/1461051582317/1479460937150/1489659893216/1499259154864/1501670702423?lk=22612&showFilter=false&hideFilter=false&disableFilter=false&filter_-1=100&filter_-1=79900&filter_-2=true'
w2 = 'https://www.alternate.de/Spielzeug/Lernen-Experimentieren/html/listings/1457363011506/1459516079131/1465308657164/1484309844578/1487070256020/1524471550366?lk=22585&showFilter=false&hideFilter=false&disableFilter=false&filter_-1=400&filter_-1=13900&filter_-2=true'
w3 = 'https://www.alternate.de/Spielzeug/Ferngesteuerte-Fahrzeuge/html/listings/1460710086507/1460710325723/1460968636916/1461321298731/1461576146133/1461576220900/1505826697456?lk=22582&showFilter=false&hideFilter=false&disableFilter=false&filter_-1=900&filter_-1=339900&filter_-2=true'
w4 = 'https://www.alternate.de/Spielzeug/Puppen-Spielfiguren/html/listings/1551447554148/1552658226942?lk=16784&showFilter=false&hideFilter=false&disableFilter=false&filter_-1=200&filter_-1=14000&filter_-2=true'
w5 = 'https://www.alternate.de/Spielzeug/Gesellschaftsspiele/html/listings/1454073847230/1454074135872/1455554585461/1455554695131/1459780263950/1460968905460/1460969231633/1460969537907/1460971476715/1461052642928/1487689963682/1487690209511/1528187238672/1528187321316/1528187387696/1534317012430/1534317729217/1542814991562?lk=16699&showFilter=false&hideFilter=false&disableFilter=false&filter_-1=100&filter_-1=23000&filter_-2=true'
w6 = 'https://www.alternate.de/Spielzeug/Geschicklichkeitsspielzeug/html/listings/1460969390893/1460970556775/1460970859655/1460971151514/1460971203369/1460971241955/1507109446319?lk=22578&showFilter=false&hideFilter=false&disableFilter=false&filter_-1=300&filter_-1=19000&filter_-2=true'
w7 = 'https://www.alternate.de/Spielzeug/Babyspielzeug/html/listings/1459507143648/1459776295875/1459776423241/1464614257041/1486541013596/1505828502926/1505988658081/1547478159883?lk=16758&showFilter=false&hideFilter=false&disableFilter=false&filter_-1=100&filter_-1=16500&filter_-2=true'
w = [w1,w2,w3,w4,w5,w6,w7]
nw = []
soup =[]
pt = []
pr = []         #initialize the lists

for i in range(len(w)):
    nw.append(requests.get(w[i]))
    nw[i].raise_for_status()
    soup.append(bs4.BeautifulSoup(nw[i].text))         # get the html of the website and put it in a soup type object

for i in range(len(soup)):
    pt.append(soup[i].select('.productLinkName'))      # let soup select the html with the product name and price
for i in range(len(soup)):
    pr.append(soup[i].select('.price'))



for i in range(len(soup)):
    for n in range(len(pt[0])):
        if i == 0:
            t = 1 + n       
        else:
            for x in range(len(soup)):
                t = (40*i)+n                # t is the rows where it should start putting the title and price
        b = pr[i][n].text[:-1]
        sheet.cell(row=t ,column=1).value = pt[i][n].text
        sheet.cell(row=t, column=2).value = b               # put the title and price in the excel file


wb.save('Price_product_webscraping.xlsx')           # Job done:), hereafter it is easy to open the excel file and sort on price to get a clean excel document

"""Written by Mohamed Taouil"""
